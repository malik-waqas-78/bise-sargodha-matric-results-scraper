import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re # Import the re module for regular expressions

def retrieve_bise_result(roll_no):
    """
    Retrieves the BISE Sargodha Matric result for a given roll number.
    It first performs a GET request to fetch the latest __VIEWSTATE and __EVENTVALIDATION
    tokens, then uses them in a POST request to get the result.

    Args:
        roll_no (str): The roll number to search for.

    Returns:
        dict or None: A dictionary containing the extracted student's result data
                      in the desired column format, or None if the request fails
                      or data cannot be parsed.
    """
    base_url = "http://119.159.230.2/biseresultday/resultday.aspx"

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }

    try:
        # Step 1: Perform a GET request to get the initial page and extract tokens
        print(f"Fetching initial page for Roll No: {roll_no} to get tokens...")
        get_response = requests.get(base_url, headers=headers)
        get_response.raise_for_status() # Raise an HTTPError for bad responses

        get_soup = BeautifulSoup(get_response.text, 'html.parser')

        # Extract __VIEWSTATE and __EVENTVALIDATION
        viewstate = get_soup.find('input', {'name': '__VIEWSTATE'})['value'] if get_soup.find('input', {'name': '__VIEWSTATE'}) else ''
        eventvalidation = get_soup.find('input', {'name': '__EVENTVALIDATION'})['value'] if get_soup.find('input', {'name': '__EVENTVALIDATION'}) else ''

        if not viewstate or not eventvalidation:
            print(f"Error: Could not find __VIEWSTATE or __EVENTVALIDATION on the initial page for {roll_no}.")
            return None

        # Step 2: Prepare the payload for the POST request with fresh tokens
        payload = {
            "__LASTFOCUS": "",
            "__EVENTTARGET": "",
            "__EVENTARGUMENT": "",
            "__VIEWSTATE": viewstate,
            "__EVENTVALIDATION": eventvalidation,
            "RbtSearchType": "Search by Roll No.",
            "TxtSearchText": roll_no,
            "BtnShowResults": "Show Result"
        }

        # Step 3: Perform the POST request to retrieve the result
        print(f"Sending POST request for Roll No: {roll_no}...")
        post_response = requests.post(base_url, data=payload, headers=headers)
        post_response.raise_for_status()  # Raise an HTTPError for bad responses (4xx or 5xx)

        soup = BeautifulSoup(post_response.text, 'html.parser')

        # Initialize student record with default empty values for all desired columns
        # Subject order changed: Computer Science and Biology first, then other Science subjects, then Arts subjects
        student_record = {
            'Roll-No': '',
            'Candidate Name': '',
            'Father Name': '',
            'Computer Science': '', # Moved to the start of science subjects
            'Biology': '',         # Moved to the start of science subjects
            'Mathematics': '',
            'Physics': '',
            'Chemistry': '',
            'Islamiyat': '',
            'Pakistan Studies': '',
            'Urdu': '',
            'English': '',
            'THQ': '', # Translation of Holy Quran
            'overall result': ''
        }

        # Extract student information and populate the student_record
        student_record['Roll-No'] = soup.find('span', id='LblRollNo').get_text(strip=True) if soup.find('span', id='LblRollNo') else ''
        student_record['Candidate Name'] = soup.find('span', id='LblName').get_text(strip=True) if soup.find('span', id='LblName') else ''
        student_record['Father Name'] = soup.find('span', id='LblFatherName').get_text(strip=True) if soup.find('span', id='LblFatherName') else ''
        student_record['overall result'] = soup.find('span', id='lblGazres').get_text(strip=True) if soup.find('span', id='lblGazres') else ''

        # Check if result data is actually present (e.g., if a valid roll number was entered)
        if not student_record['Roll-No']:
            print(f"No result found for Roll No: {roll_no}. It might be an invalid roll number or the page structure changed.")
            return None

        # Define a mapping from subject names in HTML to desired Excel column names
        subject_column_map = {
            "ISLAMIYAT (COMPULSORY)": "Islamiyat",
            "PAKISTAN STUDIES (COMPULSORY)": "Pakistan Studies",
            "URDU": "Urdu",
            "ENGLISH": "English",
            "MATHEMATICS": "Mathematics",
            "PHYSICS": "Physics",
            "CHEMISTRY": "Chemistry",
            "COMPUTER SCIENCE": "Computer Science",
            "TRANSLATION OF THE HOLY QURAN": "THQ",
            "BIOLOGY": "Biology"
        }

        # Extract subject marks
        result_table = soup.find('table', id='TblResult')
        if result_table:
            # Skip the first 5 rows which are headers/student info
            rows = result_table.find_all('tr')[5:]
            for row in rows:
                cols = row.find_all(['td', 'th'])
                if len(cols) >= 3: # Ensure at least subject name and marks obtained
                    subject_name_html = cols[0].get_text(strip=True)
                    marks_obtained_html = cols[2].get_text(strip=True)

                    # Map the HTML subject name to our desired column name
                    excel_column_name = subject_column_map.get(subject_name_html)

                    if excel_column_name:
                        student_record[excel_column_name] = marks_obtained_html
        return student_record

    except requests.exceptions.RequestException as e:
        print(f"Error during request for Roll No {roll_no}: {e}")
        return None
    except Exception as e:
        print(f"An error occurred during parsing or data extraction for Roll No {roll_no}: {e}")
        return None

def append_to_excel(data, filename="bise_results.xlsx"):
    """
    Appends a list of dictionaries (student data) to an Excel file.
    Creates the file with headers if it doesn't exist.
    Highlights failed subject cells with a light red background.
    Applies enhanced Excel formatting.

    Args:
        data (list of dict): List of dictionaries, where each dict is a student's record.
        filename (str): The name of the Excel file.
    """
    if not data:
        print("No data to append.")
        return

    # Define the desired column order explicitly (Computer Science and Biology first, then other Science subjects, then Arts subjects)
    column_order = [
        'Roll-No', 'Candidate Name', 'Father Name', 'Computer Science', 'Biology',
        'Mathematics', 'Physics', 'Chemistry', 'Islamiyat', 'Pakistan Studies', 'Urdu',
        'English', 'THQ', 'overall result'
    ]

    df_new = pd.DataFrame(data, columns=column_order)

    # Define the light red fill style
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Mapping of common failed subject abbreviations/keywords to Excel column names
    failed_subject_keywords_map = {
        "BIO": "Biology",
        "PHY": "Physics",
        "CHM": "Chemistry",
        "EGL": "English",
        "URU": "Urdu",
        "MAT": "Mathematics",
        "THQ": "THQ",
        "PKS": "Pakistan Studies",
        "ISM": "Islamiyat",
        "CSC": "Computer Science",
        "CS": "Computer Science"
    }

    if os.path.exists(filename):
        try:
            # Load the existing workbook and get the target sheet
            book = load_workbook(filename)
            sheet_name = 'BISE Sargodha Matric Results'
            
            # If the sheet does not exist, create it. Otherwise, get it.
            if sheet_name not in book.sheetnames:
                sheet = book.create_sheet(sheet_name)
            else:
                sheet = book[sheet_name]

            # Calculate the starting row for new data
            start_row = sheet.max_row

            # Use ExcelWriter with the loaded book
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay', book=book) as writer:
                # Write new data to the existing sheet, starting after the last row
                df_new.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)

            # Apply highlighting to the newly added rows using the 'sheet' object
            # Iterate through the new rows in the DataFrame to apply highlighting
            for r_idx, row_data in df_new.iterrows():
                current_excel_row = start_row + r_idx + 1 # +1 because Excel rows are 1-indexed
                overall_result_string = str(row_data.get('overall result', '')).upper() # Get overall result and convert to uppercase

                # Identify failed subjects from the overall result string
                failed_subjects = set()
                # Split the overall_result_string into words/tokens
                words = overall_result_string.replace('/', ' ').replace('-', ' ').split()
                for word in words:
                    # Start with the word, strip whitespace, and convert to uppercase
                    clean_word = word.strip()
                    
                    # Remove '(PR)' suffix
                    clean_word = clean_word.replace('(PR)', '')

                    # Use regex to remove 'II' then 'I' suffixes
                    clean_word = re.sub(r'II$', '', clean_word)
                    clean_word = re.sub(r'I$', '', clean_word)
                    
                    if clean_word in failed_subject_keywords_map:
                        failed_subjects.add(failed_subject_keywords_map[clean_word])

                # Apply highlighting based on identified failed subjects
                for col_name in column_order: # Iterate through all possible subject columns
                    if col_name in failed_subject_keywords_map.values(): # Check if it's a subject column
                        if col_name in failed_subjects:
                            # Get the column index (1-based) for the current subject column
                            col_idx = df_new.columns.get_loc(col_name) + 1
                            cell = sheet.cell(row=current_excel_row, column=col_idx)
                            cell.fill = red_fill

            # Reapply header styling, auto-adjust column widths, and freeze panes to the entire sheet
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            for col_idx in range(1, len(column_order) + 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment

            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            sheet.freeze_panes = 'C2' # Freezes row 1 and columns A, B

            # Save the workbook after all modifications
            book.save(filename)
            print(f"Data appended to '{filename}' with highlighting and formatting successfully.")

        except Exception as e:
            print(f"Error appending to existing Excel file with highlighting: {e}")
            # If an error occurs during append, try creating a new file as a fallback
            print("Attempting to create a new file instead due to append error...")
            # Fallback to creating a new file if append fails
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_new.to_excel(writer, sheet_name='BISE Sargodha Matric Results', index=False)
                sheet = writer.sheets['BISE Sargodha Matric Results']
                
                # Apply header styling
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                for col_idx in range(1, len(column_order) + 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.alignment = header_alignment

                # Auto-adjust column widths
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column].width = adjusted_width

                # Freeze panes
                sheet.freeze_panes = 'C2'

                # Apply highlighting to all rows in the new file (including the first row of data)
                for r_idx in range(2, sheet.max_row + 1): # Start from row 2 (after header)
                    row_data_from_df = df_new.iloc[r_idx - 2] # Adjust for 0-indexed DataFrame vs 1-indexed sheet
                    overall_result_string = str(row_data_from_df.get('overall result', '')).upper()
                    
                    failed_subjects = set()
                    words = overall_result_string.replace('/', ' ').replace('-', ' ').split()
                    for word in words:
                        clean_word = word.strip()
                        clean_word = clean_word.replace('(PR)', '')
                        clean_word = re.sub(r'II$', '', clean_word)
                        clean_word = re.sub(r'I$', '', clean_word)
                        if clean_word in failed_subject_keywords_map:
                            failed_subjects.add(failed_subject_keywords_map[clean_word])

                    for col_name in column_order:
                        if col_name in failed_subject_keywords_map.values():
                            if col_name in failed_subjects:
                                col_idx = df_new.columns.get_loc(col_name) + 1
                                cell = sheet.cell(row=r_idx, column=col_idx)
                                cell.fill = red_fill
            print(f"New Excel file '{filename}' created as fallback and data saved with highlighting and formatting.")

    else:
        # Create new file and apply conditional formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_new.to_excel(writer, sheet_name='BISE Sargodha Matric Results', index=False)

            sheet = writer.sheets['BISE Sargodha Matric Results']

            # Apply header styling
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            for col_idx in range(1, len(column_order) + 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment

            # Auto-adjust column widths
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            # Freeze panes
            sheet.freeze_panes = 'C2'

            # Apply conditional formatting to all rows in the new file
            # Start from row 2 (after header)
            for r_idx in range(2, sheet.max_row + 1):
                # Adjust for 0-indexed DataFrame vs 1-indexed sheet
                row_data_from_df = df_new.iloc[r_idx - 2]
                overall_result_string = str(row_data_from_df.get('overall result', '')).upper()
                
                failed_subjects = set()
                words = overall_result_string.replace('/', ' ').replace('-', ' ').split()
                for word in words:
                    clean_word = word.strip()
                    clean_word = clean_word.replace('(PR)', '')
                    clean_word = re.sub(r'II$', '', clean_word)
                    clean_word = re.sub(r'I$', '', clean_word)
                    if clean_word in failed_subject_keywords_map:
                        failed_subjects.add(failed_subject_keywords_map[clean_word])

                for col_name in column_order:
                    if col_name in failed_subject_keywords_map.values():
                        if col_name in failed_subjects:
                            col_idx = df_new.columns.get_loc(col_name) + 1
                            cell = sheet.cell(row=r_idx, column=col_idx)
                            cell.fill = red_fill
        print(f"New Excel file '{filename}' created and data saved with highlighting and formatting.")

def get_roll_number_range():
    """
    Prompts the user for starting and ending roll numbers and validates the input.

    Returns:
        tuple: A tuple containing (start_roll_no, end_roll_no) as integers.
    """
    while True:
        try:
            start_roll_no_str = input("Enter the starting roll number (e.g., 520001): ")
            end_roll_no_str = input("Enter the ending roll number (e.g., 520010): ")

            start_roll_no = int(start_roll_no_str)
            end_roll_no = int(end_roll_no_str)

            if start_roll_no <= 0 or end_roll_no <= 0:
                print("Roll numbers must be positive integers. Please try again.")
                continue
            if start_roll_no > end_roll_no:
                print("Starting roll number cannot be greater than ending roll number. Please try again.")
                continue
            return start_roll_no, end_roll_no
        except ValueError:
            print("Invalid input. Please enter valid integer roll numbers.")

def main():
    """
    Main function to orchestrate the retrieval and saving of BISE results.
    """
    start_roll_no, end_roll_no = get_roll_number_range()
    roll_numbers_to_search = [str(roll) for roll in range(start_roll_no, end_roll_no + 1)]

    all_students_results = []
    for roll_no in roll_numbers_to_search:
        print(f"Retrieving result for Roll No: {roll_no}...")
        student_result = retrieve_bise_result(roll_no)
        if student_result:
            all_students_results.append(student_result)
        else:
            print(f"Could not retrieve result for Roll No: {roll_no}")
        print("-" * 30)

    if all_students_results:
        append_to_excel(all_students_results, "bise_matric_results.xlsx")
    else:
        print("No results were successfully retrieved to save to Excel.")

if __name__ == "__main__":
    main()
